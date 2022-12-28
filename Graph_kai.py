
# coding: utf-8
import pandas as pd
import openpyxl
import matplotlib.pyplot as plt
import tkinter as tk
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg


book=openpyxl.load_workbook("/Users/hayashi/python/book.xlsx")
sheets=book.sheetnames
df=[]
for i in range(len(sheets)):
    df.append(pd.read_excel("/Users/hayashi/python/book.xlsx",sheet_name=sheets[i]))
df_tmp=df.copy()
column=[]
column=df_tmp[0].columns


def clear():
    graph_list.clear()
    for i in range(len(label_list)):
        label_list[i].destroy()

def end():
    main_win.destroy()

def select_list():
    selected_index = listbox.curselection()
    selected_module = listbox.get(selected_index)
    graph_list.append(selected_module)
    label = tk.Label(text=selected_module)
    label.pack(anchor=tk.NE)
    label_list.append(label)

def mean() :
    add_column=textBox.get()
    print(add_column)
    for i in range(len(sheets)):
        df_tmp[i][add_column]=0
        for j in range(len(selected_column)):
            df_tmp[i][add_column]+=df_tmp[i][selected_column[j]]
        df_tmp[i][add_column]=(df_tmp[i][add_column])/len(selected_column)

def excel_output():
    new_excel_name=textBox3.get()
    excel_name=new_excel_name+'.xlsx'
    writer=pd.ExcelWriter(excel_name,engin="xlswriter")
    for i in range(len(sheets)):
        df_tmp[i].to_excel(writer,sheet_name=sheets[i],index=False)
    writer.save()
    writer.close()
    wb=openpyxl.load_workbook(excel_name)
    sh=wb.create_sheet('graph')
    img=openpyxl.drawing.image.Image(fig_path)
    sh.add_image(img,'B2')
    wb.save(excel_name)
    wb.close()

def reset1():
    textBox.delete(0,tk.END)
def reset2():
    textBox2.delete(0,tk.END)
def reset3():
    textBox3.delete(0,tk.END)

def graph():
    graph_name=textBox2.get()

main_win =tk.Tk()
main_win.title("graph window ")
main_win.geometry("600x800")

canvas = tk.Canvas(main_win,width=300,height=200)

button = tk. Button(text='clear', font=("Meiryo","12", "normal"), command=clear)
button.place(relx = 0.1,rely =0.8,anchor = tk.CENTER,width = 70, height=50)

button = tk.Button(text='end', font= ("Meiryo","12", "normal"), command =end)
button.place(relx = 0.9, rely = 0.8, anchor = tk.CENTER, width = 70, height=50)

button=tk.Button(text='graph',font=("Meiryo","12","normal"),command=graph)
button.place(relx = 0.5,rely =0.8, anchor = tk.CENTER, width = 70, height=50)

button = tk.Button(text='mean', font=("Meiryo", "12", "normal"), command=mean)
button.place(relx = 0.3, rely = 0.8, anchor = tk.CENTER, width = 70, height=50)
button.bind('<1>', lambda event: listbox.insert(tk.END, add_column))

button = tk.Button(text='Excel出力', font=("Meiryo", "12", "normal"), command=excel_output)
button.place(relx = 0.7, rely = 0.8, anchor = tk.CENTER, width = 70, height=50)


list=[]
selected_column=[]
label_list=[]
graph_list=[]

for i in range(len(column)):
    list.append(column[i])
list_var=tk.StringVar(value=list)
listbox =tk.Listbox(main_win,width=40,height=20,listvariable=list_var)
listbox.grid(row=0, column=1, padx=10, pady=10)
listbox.bind('<<ListboxSelect>>', lambda e: select_list())
listbox.pack()

add_column_label = tk.Label(text="追加する列名")
add_column_label.place(x=30,y=400)

textBox= tk.Entry(width=20)
textBox.place(x=30,y=420)

button = tk.Button(text='clear', font=("Meiryo","12", "normal"), command = reset1)
button.place(x=180, y=445,width= 40, height = 20)

image_label=tk.Label(text="新規エクセルシート名")
image_label.place(x=30,y=460)

textBox2= tk.Entry(width=20)
textBox2.place(x=30,y=480)

button = tk.Button(text='clear', font=("Meiryo", "12", "normal"), command = reset2)
button.place(x=180, y=505,width = 40, height = 20)

exel_label = tk.Label(text="保存する画像名")
exel_label.place(x=30,y=520)

textBox3=tk.Entry(width=20)
textBox3.place(x=30,y=540)

button = tk.Button(text='clear', font=("Meiryo", "12", "normal"), command = reset3)
button.place(x=180, y=565,width=40, height = 20)

for j in range(len(selected_column)):
    figure,ax = plt.subplots()
    for i in range (len(sheets)):
        ax.plot(df_tmp[i][selected_column[j]], label="case" +str(i+1))
    ax.set_xlabel('経過時間 (s)',fontname="MS Gothic")
    ax.set_ylabel('温度 (℃)',fontname="MS Gothic")
    plt.grid()
    ax.legend(loc= (1.05, 0.7))
    pit.title(selected_column[j], fontname="MS Gothic")
    figure.subplots_adjust(right=0.8)
    figure.show()
    fig_path="/Users/hayashi/python/graph"+graph_name+"png"
    fig.savefig(fig_path)

main_win.mainloop()
